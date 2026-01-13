from __future__ import annotations

from pathlib import Path
import pandas as pd
import numpy as np
import xgboost as xgb
import math
import shutil

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# -----------------------
# Paths
# -----------------------
NBA_LOGS = Path("data/raw/player_game_logs.csv")
PROPS_NORM = Path("data/odds_logs/assists_props_normalized.csv")

# Models
MIN_MODEL_PATH = Path("models/minutes_xgb.json")  # required
TEAM_MAP_PATH = Path("data/processed/player_team_map_current.csv")  # required

# Output
OUT_PATH = Path("data/processed/today_assists_prop_predictions.xlsx")

# Books
BOOK_PRIORITY = ["fanduel", "bet365"]


# -----------------------
# Team mapping (Odds API full names -> abbreviations)
# -----------------------
TEAM_NAME_TO_ABBR = {
    "atlanta hawks": "ATL",
    "boston celtics": "BOS",
    "brooklyn nets": "BKN",
    "charlotte hornets": "CHA",
    "chicago bulls": "CHI",
    "cleveland cavaliers": "CLE",
    "dallas mavericks": "DAL",
    "denver nuggets": "DEN",
    "detroit pistons": "DET",
    "golden state warriors": "GSW",
    "houston rockets": "HOU",
    "indiana pacers": "IND",
    "los angeles clippers": "LAC",
    "la clippers": "LAC",
    "los angeles lakers": "LAL",
    "la lakers": "LAL",
    "memphis grizzlies": "MEM",
    "miami heat": "MIA",
    "milwaukee bucks": "MIL",
    "minnesota timberwolves": "MIN",
    "new orleans pelicans": "NOP",
    "new york knicks": "NYK",
    "oklahoma city thunder": "OKC",
    "orlando magic": "ORL",
    "philadelphia 76ers": "PHI",
    "phoenix suns": "PHX",
    "portland trail blazers": "POR",
    "sacramento kings": "SAC",
    "san antonio spurs": "SAS",
    "toronto raptors": "TOR",
    "utah jazz": "UTA",
    "washington wizards": "WAS",
}


# -----------------------
# Helpers
# -----------------------
def archive_run_assists(
    out_xlsx_path: Path,
    props_norm_path: Path,
    overs_tbl: pd.DataFrame,
    unders_tbl: pd.DataFrame,
    tz: str = "America/Toronto",
) -> Path:
    """
    Archives today's assists run output + exact input props snapshot + picks table.

    Creates:
      data/archives/assists/YYYY-MM-DD/
        - assists_predictions.xlsx
        - assists_props_normalized.csv
        - picks.csv
    """
    run_date = pd.Timestamp.now(tz=tz).date().isoformat()
    arch_dir = Path("data/archives/assists") / run_date
    arch_dir.mkdir(parents=True, exist_ok=True)

    # 1) Copy final Excel output
    xlsx_dst = arch_dir / "assists_predictions.xlsx"
    if out_xlsx_path.exists():
        shutil.copy2(out_xlsx_path, xlsx_dst)

    # 2) Copy the props snapshot used
    props_dst = arch_dir / "assists_props_normalized.csv"
    if props_norm_path.exists():
        shutil.copy2(props_norm_path, props_dst)

    # 3) Save picks.csv (combine overs + unders, keep section label)
    picks = []
    if overs_tbl is not None and not overs_tbl.empty:
        o = overs_tbl.copy()
        o.insert(0, "SECTION", "OVER")
        picks.append(o)
    if unders_tbl is not None and not unders_tbl.empty:
        u = unders_tbl.copy()
        u.insert(0, "SECTION", "UNDER")
        picks.append(u)

    if picks:
        picks_df = pd.concat(picks, ignore_index=True)
        picks_df.to_csv(arch_dir / "picks.csv", index=False)

    print(f"📦 Archived assists run to: {arch_dir.resolve()}")
    return arch_dir

def norm_name(s: str) -> str:
    s = str(s).lower().strip()
    for tok in [" jr.", " sr.", " iii", " ii", " iv"]:
        s = s.replace(tok, "")
    s = s.replace(".", "").replace("'", "")
    s = " ".join(s.split())
    return s


def american_to_implied(odds) -> float:
    if odds is None or (isinstance(odds, float) and np.isnan(odds)):
        return float("nan")
    odds = float(odds)
    if odds > 0:
        return 100.0 / (odds + 100.0)
    return (-odds) / ((-odds) + 100.0)


def norm_cdf(x: float) -> float:
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def load_booster(path: Path) -> xgb.Booster:
    if not path.exists():
        raise RuntimeError(f"Missing model file: {path}")
    b = xgb.Booster()
    b.load_model(str(path))
    return b


def team_to_abbr(team_str: str) -> str | None:
    """
    Converts Odds API team name to NBA abbreviation if possible.
    If already an abbreviation-like string, returns upper.
    """
    if team_str is None or (isinstance(team_str, float) and np.isnan(team_str)):
        return None
    t = str(team_str).strip()
    if len(t) <= 4 and t.isalpha():
        return t.upper()
    key = t.lower().strip()
    return TEAM_NAME_TO_ABBR.get(key)


# -----------------------
# Feature builders (historical data)
# -----------------------
def build_latest_player_features(nba: pd.DataFrame) -> pd.DataFrame:
    """
    One row per player from historical logs, using rolling features (shift(1) to avoid leakage).
    Also builds assists rate + volatility for the assists model.
    """
    nba = nba.copy()
    nba["GAME_DATE"] = pd.to_datetime(nba["GAME_DATE"])
    nba = nba.sort_values(["PLAYER_ID", "GAME_DATE"])

    # Ensure numeric
    for c in ["MIN", "PTS", "AST", "FGA", "FTA", "FG3A", "TOV", "REB"]:
        if c in nba.columns:
            nba[c] = pd.to_numeric(nba[c], errors="coerce")

    if "AST" not in nba.columns:
        raise RuntimeError("NBA logs file is missing AST column. Update/fetch logs to include assists.")

    rows = []
    for _, pdf in nba.groupby("PLAYER_ID"):
        pdf = pdf.sort_values("GAME_DATE")

        # Minutes features (for minutes model input)
        pdf["min_last5"] = pdf["MIN"].shift(1).rolling(5).mean()
        pdf["min_last10"] = pdf["MIN"].shift(1).rolling(10).mean()
        pdf["min_std_last10"] = pdf["MIN"].shift(1).rolling(10).std()

        # These are part of minutes model input in your current setup
        pdf["pts_last10"] = pdf["PTS"].shift(1).rolling(10).mean()
        pdf["fga_last10"] = pdf["FGA"].shift(1).rolling(10).mean()
        pdf["fta_last10"] = pdf["FTA"].shift(1).rolling(10).mean()
        pdf["fg3a_last10"] = pdf["FG3A"].shift(1).rolling(10).mean()
        pdf["tov_last10"] = pdf["TOV"].shift(1).rolling(10).mean()
        pdf["reb_last10"] = pdf["REB"].shift(1).rolling(10).mean()

        # Assists model features
        pdf["ast_last10"] = pdf["AST"].shift(1).rolling(10).mean()

        pdf["ast_per_min_last10"] = (
            pdf["AST"].shift(1).rolling(10).sum()
            / pdf["MIN"].shift(1).rolling(10).sum()
        )

        pdf["ast_std_last10"] = pdf["AST"].shift(1).rolling(10).std()

        # Context
        pdf["rest_days"] = (pdf["GAME_DATE"] - pdf["GAME_DATE"].shift(1)).dt.days
        pdf["is_b2b"] = (pdf["rest_days"] == 1).astype(int)
        pdf["is_home"] = pdf["MATCHUP"].astype(str).str.contains(" vs. ").astype(int)

        if "START_POSITION" in pdf.columns:
            pdf["is_starter"] = (pdf["START_POSITION"].astype(str).str.strip() != "").astype(int)
        else:
            pdf["is_starter"] = (pdf["MIN"] >= 24).astype(int)

        rows.append(pdf.iloc[-1:].copy())

    f = pd.concat(rows, ignore_index=True)
    f["player_norm"] = f["PLAYER_NAME"].apply(norm_name)

    keep = [
        "PLAYER_ID", "PLAYER_NAME", "player_norm",

        # minutes features
        "min_last5", "min_last10", "min_std_last10",
        "pts_last10", "fga_last10", "fta_last10", "fg3a_last10",
        "tov_last10", "reb_last10",
        "rest_days", "is_b2b", "is_home", "is_starter",

        # assists features
        "ast_last10", "ast_per_min_last10", "ast_std_last10",
    ]
    keep = [c for c in keep if c in f.columns]
    return f[keep].drop_duplicates(subset=["player_norm"], keep="last")


# -----------------------
# Excel formatting helpers
# -----------------------
def write_section(ws, start_row: int, title: str, df_table: pd.DataFrame,
                  bar_fill, header_fill, row_fill) -> int:
    cols = ["PLAYER NAME", "PLAYER TEAM", "OPPONENT TEAM", "PROP", "ODDS", "AI RATING"]
    ncols = len(cols)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Section bar row
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=ncols)
    cell = ws.cell(row=start_row, column=1, value=title)
    cell.fill = bar_fill
    cell.font = Font(bold=True, color="000000", size=12)
    cell.alignment = center

    for c in range(1, ncols + 1):
        cc = ws.cell(row=start_row, column=c)
        cc.fill = bar_fill
        cc.border = border
        cc.alignment = center

    # Header row
    hdr_row = start_row + 1
    for i, name in enumerate(cols, start=1):
        ws.cell(row=hdr_row, column=i, value=name)

    for c in range(1, ncols + 1):
        cc = ws.cell(row=hdr_row, column=c)
        cc.fill = header_fill
        cc.font = Font(bold=True, color="000000")
        cc.border = border
        cc.alignment = center

    # Data rows
    cur = hdr_row + 1
    for _, r in df_table.iterrows():
        ws.cell(row=cur, column=1, value=str(r["PLAYER NAME"]))
        ws.cell(row=cur, column=2, value=str(r["PLAYER TEAM"]))
        ws.cell(row=cur, column=3, value=str(r["OPPONENT TEAM"]))
        ws.cell(row=cur, column=4, value=str(r["PROP"]))
        ws.cell(row=cur, column=5, value=int(r["ODDS"]))
        ws.cell(row=cur, column=6, value=float(r["AI RATING"]))

        for c in range(1, ncols + 1):
            cc = ws.cell(row=cur, column=c)
            cc.fill = row_fill
            cc.border = border
            cc.alignment = center
        cur += 1

    return cur  # COMPACT: no extra blank row


# -----------------------
# Main
# -----------------------
def main():
    if not NBA_LOGS.exists():
        raise RuntimeError(f"Missing NBA logs file: {NBA_LOGS}")
    if not PROPS_NORM.exists():
        raise RuntimeError(f"Missing normalized props file: {PROPS_NORM}")
    if not MIN_MODEL_PATH.exists():
        raise RuntimeError(f"Missing minutes model: {MIN_MODEL_PATH}")
    if not TEAM_MAP_PATH.exists():
        raise RuntimeError(
            f"Missing current team map: {TEAM_MAP_PATH}\n"
            "Run: python data/raw/fetch_current_player_teams.py"
        )

    nba = pd.read_csv(NBA_LOGS)
    props = pd.read_csv(PROPS_NORM)
    
    props["commence_time"] = pd.to_datetime(props["commence_time"], utc=True, errors="coerce")
    local_tz = "America/Toronto"
    props_local_date = props["commence_time"].dt.tz_convert(local_tz).dt.date
    today_local = pd.Timestamp.now(tz=local_tz).date()
    props = props[props_local_date == today_local].copy()

    if props.empty:
        raise RuntimeError(f"No assists props after filtering to local date {today_local}.")

    # Guard against stale normalized file
    most_recent = props["commence_time"].max()
    most_recent_local = most_recent.tz_convert("America/Toronto").date()
    if most_recent_local != pd.Timestamp.now(tz="America/Toronto").date():
        raise RuntimeError(f"Normalized assists file looks stale (latest game date {most_recent_local}). Re-fetch/normalize.")


    # Filter books and pick best (FD > bet365)
    props["book_key"] = props["book_key"].astype(str).str.lower().str.strip()
    props = props[props["book_key"].isin(BOOK_PRIORITY)].copy()
    if props.empty:
        raise RuntimeError("No props left after filtering to FanDuel/bet365.")

    props["book_rank"] = props["book_key"].apply(lambda b: BOOK_PRIORITY.index(b))
    props = props.sort_values(["commence_time", "event_id", "player", "line", "book_rank"])
    props = props.drop_duplicates(subset=["event_id", "player", "line"], keep="first")

    props["player_norm"] = props["player"].apply(norm_name)
    props["home_abbr"] = props["home_team"].apply(team_to_abbr)
    props["away_abbr"] = props["away_team"].apply(team_to_abbr)

    # Load current team map (today roster)
    team_map = pd.read_csv(TEAM_MAP_PATH)
    if "player_norm" not in team_map.columns:
        team_map["player_norm"] = team_map["PLAYER_NAME"].apply(norm_name)

    if "TEAM_ABBREVIATION" not in team_map.columns:
        raise RuntimeError("TEAM_MAP file missing TEAM_ABBREVIATION column.")

    team_map["TEAM_ABBREVIATION"] = team_map["TEAM_ABBREVIATION"].astype(str).str.upper().str.strip()

    df = props.merge(team_map[["player_norm", "TEAM_ABBREVIATION"]], on="player_norm", how="left")
    df = df.rename(columns={"TEAM_ABBREVIATION": "player_team_abbr"})

    # Infer opponent abbreviation using CURRENT team + matchup
    def infer_opp_abbr(row):
        pt = row.get("player_team_abbr")
        ha = row.get("home_abbr")
        aa = row.get("away_abbr")
        if pd.isna(pt) or pt is None:
            return None
        if pt == ha:
            return aa
        if pt == aa:
            return ha
        return None

    df["opp_team_abbr"] = df.apply(infer_opp_abbr, axis=1)

    # Player historical features (includes assists features now)
    player_feats = build_latest_player_features(nba)
    df = df.merge(player_feats, on="player_norm", how="left")

    # Implied probs from odds
    df["p_over_implied"] = df["odds_over"].apply(american_to_implied)
    df["p_under_implied"] = df["odds_under"].apply(american_to_implied)

    # Minutes prediction (same model)
    min_booster = load_booster(MIN_MODEL_PATH)

    MIN_FEATURES = [
        "min_last5", "min_last10", "min_std_last10",
        "pts_last10", "fga_last10", "fta_last10", "fg3a_last10",
        "tov_last10", "reb_last10",
        "rest_days", "is_b2b", "is_home",
        "is_starter",
    ]

    REQUIRED = MIN_FEATURES + [
        "ast_per_min_last10",
        "ast_std_last10",
        "line",
        "p_over_implied", "p_under_implied",
        "player_team_abbr", "opp_team_abbr",
        "home_abbr", "away_abbr",
    ]

    pred_df = df.dropna(subset=REQUIRED).copy()
    if pred_df.empty:
        print("❌ No rows left after merges.")
        print("Likely causes: missing players in team map, name mismatch, or team mapping issues.")
        return

    dmin = xgb.DMatrix(pred_df[MIN_FEATURES].values, feature_names=MIN_FEATURES)
    pred_df["pred_minutes"] = np.clip(min_booster.predict(dmin), 0, 42)

    # Assists mean = minutes * assists-per-minute rate
    pred_df["ast_mean"] = pred_df["pred_minutes"] * pred_df["ast_per_min_last10"]

    # Normal uncertainty for assists (smaller scale than points)
    pred_df["sigma"] = pred_df["ast_std_last10"].astype(float).clip(lower=1.5).fillna(2.0)

    # P(Over)
    z = (pred_df["line"] - pred_df["ast_mean"]) / pred_df["sigma"]
    pred_df["p_over_model"] = 1.0 - z.apply(norm_cdf)
    pred_df["p_under_model"] = 1.0 - pred_df["p_over_model"]

    # Edge Over/Under
    pred_df["edge_over"] = pred_df["p_over_model"] - pred_df["p_over_implied"]
    pred_df["edge_under"] = pred_df["p_under_model"] - pred_df["p_under_implied"]

    # Best side per prop row
    pred_df["best_side"] = np.where(pred_df["edge_over"] >= pred_df["edge_under"], "OVER", "UNDER")
    pred_df["best_edge"] = np.where(pred_df["best_side"] == "OVER", pred_df["edge_over"], pred_df["edge_under"])
    pred_df["best_odds"] = np.where(pred_df["best_side"] == "OVER", pred_df["odds_over"], pred_df["odds_under"])

    # AI rating = edge * 100
    pred_df["ai_rating"] = (pred_df["best_edge"] * 100).round(1)

    pred_df = pred_df.replace([np.inf, -np.inf], np.nan).dropna(subset=["ai_rating"])

    # -----------------------
    # Selection rules:
    # 1) Max 1 pick per player total
    # 2) Top 11 overall
    # 3) Max 5 unders (no minimum)
    # -----------------------
    ranked = (
        pred_df.sort_values("ai_rating", ascending=False)
        .drop_duplicates(subset=["player_norm", "line"], keep="first")  # avoid duplicate same line
        .drop_duplicates(subset=["player_norm"], keep="first")         # MAX 1 pick per player total
        .copy()
    )

    unders = ranked[ranked["best_side"] == "UNDER"].head(5).copy()     # MAX 5 UNDERS
    overs = ranked[ranked["best_side"] == "OVER"].copy()

    remaining = 11 - len(unders)
    overs = overs.head(max(0, remaining)).copy()

    def build_display(df_part: pd.DataFrame) -> pd.DataFrame:
        return pd.DataFrame({
            "PLAYER NAME": df_part["player"].astype(str),
            "PLAYER TEAM": df_part["player_team_abbr"].astype(str),
            "OPPONENT TEAM": df_part["opp_team_abbr"].astype(str),
            "PROP": df_part.apply(lambda r: f"{r['best_side']} {r['line']} ASSISTS", axis=1),
            "ODDS": df_part["best_odds"].astype(int),
            "AI RATING": df_part["ai_rating"].astype(float),
        })

    overs_tbl = build_display(overs)
    unders_tbl = build_display(unders)

    # -----------------------
    # Write formatted Excel
    # -----------------------
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        sheet_name = "Props"
        pd.DataFrame().to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]

        cols = ["PLAYER NAME", "PLAYER TEAM", "OPPONENT TEAM", "PROP", "ODDS", "AI RATING"]
        ncols = len(cols)

        # Colors (format only)
        title_fill = PatternFill("solid", fgColor="3A3A3A")

        overs_bar_fill = PatternFill("solid", fgColor="6BCB63")
        overs_header_fill = PatternFill("solid", fgColor="BFE8B9")
        overs_row_fill = PatternFill("solid", fgColor="E9F6E7")

        unders_bar_fill = PatternFill("solid", fgColor="E06A5F")
        unders_header_fill = PatternFill("solid", fgColor="F3B3AD")
        unders_row_fill = PatternFill("solid", fgColor="F9D7D4")

        # Title row
        title = f"@Jayssportsanalytics - NBA Player Assists Model - {pd.Timestamp.now().strftime('%m/%d/%Y')}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        tcell = ws.cell(row=1, column=1, value=title)
        tcell.fill = title_fill
        tcell.font = Font(bold=True, color="FFFFFF", size=14)
        tcell.alignment = Alignment(horizontal="center", vertical="center")

        # COMPACT layout
        next_row = 2
        next_row = write_section(
            ws=ws,
            start_row=next_row,
            title="OVERS",
            df_table=overs_tbl,
            bar_fill=overs_bar_fill,
            header_fill=overs_header_fill,
            row_fill=overs_row_fill,
        )

        next_row = write_section(
            ws=ws,
            start_row=next_row,
            title="UNDERS",
            df_table=unders_tbl,
            bar_fill=unders_bar_fill,
            header_fill=unders_header_fill,
            row_fill=unders_row_fill,
        )

        # Column widths
        widths = [24, 14, 16, 28, 10, 12]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"

    print(f"✅ Saved Top-11 mixed (max 5 unders, max 1 pick per player) to: {OUT_PATH.resolve()}")
    print("OVERS:", len(overs_tbl), "| UNDERS:", len(unders_tbl))
    
    # Auto-archive today's assists run
    archive_run_assists(
        out_xlsx_path=OUT_PATH,
        props_norm_path=PROPS_NORM,
        overs_tbl=overs_tbl,
        unders_tbl=unders_tbl,
    )

if __name__ == "__main__":
    main()
